#!/usr/bin/env python3
import json,sys,urllib.request
from pathlib import Path

URL='https://assets.publishing.service.gov.uk/media/60538b91e90e07527df82ae4/McCance_Widdowsons_Composition_of_Foods_Integrated_Dataset_2021..xlsx'
out=Path(sys.argv[1] if len(sys.argv)>1 else '/tmp/cofid-index.json')
xlsx=Path('/tmp/cofid-2021.xlsx')
if not xlsx.exists(): urllib.request.urlretrieve(URL,xlsx)
try:
 import openpyxl
except ImportError:
 raise SystemExit('openpyxl required')
wb=openpyxl.load_workbook(xlsx,read_only=True,data_only=True)

def norm(x): return ' '.join(str(x or '').strip().lower().split())
def num(x):
 try:
  if x in (None,'','N','Tr','trace','-'): return 0.0
  return float(x)
 except: return None

wanted={'food code':'code','food name':'name','protein (g)':'protein_g','fat (g)':'fat_g','carbohydrate (g)':'carbohydrate_g','energy (kcal) (kcal)':'kcal','aoac fibre (g)':'fibre_g'}
rows=[]; source_sheet=None
for ws in wb.worksheets:
 header=None; idx={}
 for rno,row in enumerate(ws.iter_rows(min_row=1,max_row=30,values_only=True),1):
  vals=[norm(v) for v in row]
  if 'food code' in vals and 'food name' in vals and any('protein' in v for v in vals):
   header=rno
   for i,v in enumerate(vals):
    if v in wanted: idx[wanted[v]]=i
   break
 if not header or not all(k in idx for k in ['code','name','kcal','protein_g','fat_g','carbohydrate_g','fibre_g']): continue
 source_sheet=ws.title
 for row in ws.iter_rows(min_row=header+1,values_only=True):
  name=row[idx['name']] if idx['name']<len(row) else None
  code=row[idx['code']] if idx['code']<len(row) else None
  if not name or not code: continue
  item={'code':str(code).strip(),'name':str(name).strip()}
  ok=True
  for k in ['kcal','protein_g','fat_g','carbohydrate_g','fibre_g']:
   v=num(row[idx[k]] if idx[k]<len(row) else None)
   if v is None: ok=False; break
   item[k]=v
  if ok: rows.append(item)
 break
if not rows: raise SystemExit('Could not locate CoFID proximates worksheet/headers')
out.write_text(json.dumps({'dataset':'CoFID 2021','source_url':URL,'sheet':source_sheet,'foods':rows},separators=(',',':')))
print(json.dumps({'dataset':'CoFID 2021','sheet':source_sheet,'foods':len(rows),'output':str(out)}))
